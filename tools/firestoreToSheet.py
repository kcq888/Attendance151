from __future__ import print_function
import os
from datetime import datetime
import pytz    # $ pip install pytz
import tzlocal # $ pip install tzlocal
# google sheet
import pickle
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
# Firebase Imports
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
from openpyxl import Workbook
from openpyxl.styles import Font

class FirestoreToSheet():
    Season = "Season2022-2023"
    Members = "members"
    AppConfig = "AppConfig"
    SignIn = "SignIn"
    SignOut = "SignOut"
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets', "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
    SAMPLE_SPREADSHEET_ID = '1aPZODds6OEt_uinIIGFDINqfkJbxqFuf3Pn_08Hd5WA'
    Logdate = "Log1052022"

    def __init__(self):
        # Firestore setup
        self.fcred = credentials.Certificate(os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"))
        firebase_admin.initialize_app(self.fcred)
        self.db = firestore.client()

        # google sheet API setup
        self.gscreds = None
        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                self.gscreds = pickle.load(token)
        self.service = build('sheets', 'v4', credentials=self.gscreds)
        # If there are no (valid) credentials available, let the user log in.
        if not self.gscreds or not self.gscreds.valid:
            if self.gscreds and self.gscreds.expired and self.gscreds.refresh_token:
                self.gscreds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    'credentials.json', self.SCOPES)
                self.gscreds = flow.run_local_server(port=0)
            # Save the credentials for the next run
            with open('token.pickle', 'wb') as token:
                pickle.dump(self.gscreds, token)
        self.index = 2
        self.name = 'A'
        self.signin = 'B'
        self.signout = 'C'
        self.value_input_option = 'USER_ENTERED'
        self.insert_data_option = 'OVERWRITE'

    def utcToLocal(self, utctime):
        local_timezone = tzlocal.get_localzone()
        return utctime.replace(tzinfo=pytz.utc).astimezone(local_timezone)

    def appendGoogleSheet(self, range, value):
            value_range_body = {
                "range": range,
                "values": [
                    [ value ]
                ]
            }
            request = self.service.spreadsheets().values().append(spreadsheetId=self.SAMPLE_SPREADSHEET_ID, range=range, valueInputOption=self.value_input_option, insertDataOption=self.insert_data_option, body=value_range_body)
            response = request.execute()

    def getAttendHistory(self, name, docref, ws):
         attnHistory = docref.collection('AttnHistory')
         if attnHistory is not None:
             docs = attnHistory.list_documents()
             for doc in docs:
                print(doc.id)
                signinout = doc.get().to_dict()
                for key in signinout.keys():
                    signin = True;
                    signout = True;
                    singindatetime = None
                    signoutdatetime = None
                    if (key == self.Logdate):
                        try:
                            signintime = signinout[self.Logdate][self.SignIn]
                            signintime = self.utcToLocal(signintime)
                            rangeName = self.Logdate + '!' + self.name + str(self.index) + ':' + self.name + str(self.index)
                            self.appendGoogleSheet(rangeName, name)
                            print(signinout[self.Logdate][self.SignIn])
                            rangeSignin = self.Logdate + '!' + self.signin + str(self.index) + ':' + self.signin + str(self.index)
                            singindatetime = signintime.strftime("%m/%d/%Y, %H:%M:%S")
                            # self.appendGoogleSheet(rangeSignin, signintime.strftime("%m/%d/%Y, %H:%M:%S"))
                        except KeyError:
                            print('Signin not found!')
                            signin = False
                        try:
                            signouttime = signinout[self.Logdate][self.SignOut]
                            signouttime = self.utcToLocal(signouttime)
                            print(signinout[self.Logdate][self.SignOut])
                            rangeSignout = self.Logdate + '!' + self.signout + str(self.index) + ':' + self.signout + str(self.index)
                            signoutdatetime = signouttime.strftime("%m/%d/%Y, %H:%M:%S")
                            # self.appendGoogleSheet(rangeSignout, signouttime.strftime("%m/%d/%Y, %H:%M:%S"))
                        except KeyError:
                            print('Signout not found!')
                            signout = False;
                        if (signin == True or signout == True):
                            self.index = self.index + 1
                        ws.append([name, singindatetime, signoutdatetime])

    def getAttendants(self):
        wb = self.setupWorkbook()
        ws = wb.active
        attendants = self.db.collection(self.Season).list_documents()
        for attendant in attendants:
            doc = attendant.get()
            try:
                name = doc.get(u'First') + ' ' + doc.get(u'Last')
                print(name)
            except KeyError:
                print('AppConfig document reached')
            self.getAttendHistory(name, attendant, ws)
        attendantLog = os.getcwd() + "/Attendant.xlsx"
        wb.save(attendantLog)

    def setupWorkbook(self):
        print("setupWorkbook")
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendant"
        ws.append(["Name", "sign In", "Sign Out"])
        for cell in ws[1]:  # loop through all cells and set the font to bold
            cell.font = Font(bold=True)
        return wb

if __name__ == "__main__":
    db = FirestoreToSheet()
    db.getAttendants()